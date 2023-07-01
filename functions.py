
import uuid
import openpyxl as xl
import random
import path 
import openai
from moviepy.editor import *
import textwrap
from gtts import gTTS

openai.api_key = path.openai_api_key



def uniq_id():
    id = str(uuid.uuid4()).replace('-', '')
    # print(id)
    return id



def topic_pick():
    # read data to the excel file using python openpyxl
    
    workbook = xl.load_workbook(path.data_file)
    workbook_active = workbook.active
    # print(workbook_active.max_column)
    # print(workbook_active.max_row) # data in between 1, 3

    index_of_row = random.randint(1, workbook_active.max_row)
    # print(index_of_row)

    topic = workbook_active.cell(row= index_of_row, column= 1 ).value
    
    return index_of_row, topic


def chatGPT(topic):
    
    chat_completion = openai.ChatCompletion.create(model="gpt-3.5-turbo", 
                                               messages=[{"role": "user", 
                                                          "content": f"explain {topic}"}])


    result = chat_completion['choices'][0]['message']['content']
    
    result = result.replace('\n\n', '\n')

    final_lines = []
    for line in result.splitlines():
        if '. ' in line:
            new_line = line.split('. ', 1)[1]
            final_lines.append(new_line)
        else:
            final_lines.append(line)

    final_lines
    
    return final_lines


def bg_video():
    video_path = path.bg_videos_path + random.choice(os.listdir(path.bg_videos_path))
    video = VideoFileClip(video_path)
    return video

def bg_image():
    image_path = path.bg_images_path + random.choice(os.listdir(path.bg_images_path))
    image_video = ImageClip(image_path)
    return image_video

def clips_bg_image():
    image_path = path.clips_bg_images + random.choice(os.listdir(path.clips_bg_images))
    image_video = ImageClip(image_path)
    return image_video

def title_bg_audio():
    audio_path = path.title_musiq + random.choice(os.listdir(path.title_musiq))
    audio = AudioFileClip(audio_path)
    return audio

def gtts_audio(txt, id):
    tts = gTTS(txt)
    file_name = 'output/' + id + '.mp3'
    tts.save(file_name)
    
    return file_name

def title_intro(topic:str, id):
    
    text = '\n'.join(textwrap.wrap(topic.upper(), width=30))
    
    video = bg_video()
    
    font_color = random.choice(path.color_codes)
    font_stroke_color = random.choice(path.stroke_colors)
    
    txt_clip = TextClip(txt=text, 
                        fontsize= 130, 
                        font='melton',
                        color= font_color, 
                        stroke_color= font_stroke_color, 
                        stroke_width=2).set_position('center', 'center')
    
    txt_clip1 = TextClip(txt=text, 
                        fontsize= 45, 
                        font='melton',
                        color= random.choice(path.color_codes), 
                        stroke_color= font_stroke_color,  
                        stroke_width=2).set_duration(1).set_position('center', 'center')
    
    txt_clip2 = TextClip(txt=text, 
                        fontsize= 90, 
                        font='melton',
                        color= random.choice(path.color_codes), 
                        stroke_color= font_stroke_color,  
                        stroke_width=2).set_duration(1).set_position('center', 'center')
    
    txt_clip3 = TextClip(txt=text, 
                        fontsize= 130, 
                        font='melton',
                        color= random.choice(path.color_codes), 
                        stroke_color= font_stroke_color,  
                        stroke_width=2).set_duration(1).set_position('center', 'center')
    
    def get_txt_color(t):
        if t < 0.5:
            return random.choice(path.color_codes)
        elif t < 1:
            return '#a32372'
        elif t < 1.5:
            return random.choice(path.color_codes)
        elif t < 2:
            return '#008080'
        elif t < 2.5:
            return random.choice(path.color_codes)
        elif t < 3:
            return 'red'
        else:
            return random.choice(path.color_codes)
    
    def modified_text_clip(get_frame, t):
        # print(t)
        modified_clip = TextClip(txt=text, 
                        fontsize= 130, 
                        font='melton',
                        color= get_txt_color(t), 
                        stroke_color= font_stroke_color, 
                        stroke_width=2).set_duration(1).set_position('center', 'center')
        
        return modified_clip.get_frame(t)    
    
    modified_txt_clip = txt_clip.fl(modified_text_clip)
    
    topic_audio_file = gtts_audio(txt=topic, id=id + '_title')
    topic_audio_clip = AudioFileClip(topic_audio_file)

    modified_txt_clip = modified_txt_clip.set_audio(topic_audio_clip)
    
    duration = 3 + topic_audio_clip.duration + 1
    
    bg_audio = title_bg_audio().set_duration(3)
    video = video.set_duration(duration).set_audio(bg_audio)
    final_video = CompositeVideoClip([video, txt_clip1, txt_clip2.set_start(1), txt_clip3.set_start(2), modified_txt_clip.set_start(3)]).set_duration(duration)
    
    # final_video.write_videofile('output/title_intro.mp4', fps=30, codec = 'libx264', logger= None)
    
    return final_video


def topic_clips(text, id, i):    
    
    transition_effect = VideoFileClip("gallery\\transition_video.mp4").subclip(5,7).resize((1920, 1080)).set_opacity(0.7)
      
    bg_video = clips_bg_image()
    
    txt = '\n'.join(textwrap.wrap(text, width=50))
    
    font_color = random.choice(path.color_codes)
    font_stroke_color = random.choice(path.stroke_colors)
    
    txt_clip = TextClip(txt=txt, 
                        fontsize= 80, 
                        color= font_color, 
                        stroke_color= font_stroke_color, 
                        stroke_width=2).set_position(lambda t : ('center', 950 - 30 *t))
    
    clip_audio_file = gtts_audio(txt=text, id=id + f'_clip_{i}')
    
    clip_audio = AudioFileClip(clip_audio_file)
    
    txt_clip = txt_clip.set_audio(clip_audio)
    
    duration = 2 + clip_audio.duration + 1
    bg_video = bg_video.set_duration(duration)
    
    final_video = CompositeVideoClip([bg_video, txt_clip.set_start(2), transition_effect]).set_duration(duration)
    
    # final_video.write_videofile('output/topic_clip.mp4', fps=30, codec = 'libx264', logger= None)
    
    return final_video


def data_manger(topic, index):
    
    #------data Deletion from Main Data file
    
    workbook = xl.load_workbook(path.data_file)
    workbook_active = workbook.active
    workbook_active.delete_rows(index, 1)
    workbook.save(path.data_file)
    workbook.close()
    
    #------- backup Finished Data 
    
    backup_workbook = xl.load_workbook(path.completed_data_file)
    backup_workbook_active = backup_workbook.active
    backup_workbook_active.append([topic])
    backup_workbook.save(path.completed_data_file)
    
    
    
    
    

